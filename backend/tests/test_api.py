"""End-to-end API tests covering upload, validation and Excel export."""

from __future__ import annotations

import io

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.main import app

XSD = """<?xml version="1.0"?>
<xs:schema xmlns:xs="http://www.w3.org/2001/XMLSchema">
  <xs:element name="person">
    <xs:complexType>
      <xs:sequence>
        <xs:element name="name" type="xs:string"/>
        <xs:element name="age" type="xs:integer"/>
      </xs:sequence>
    </xs:complexType>
  </xs:element>
</xs:schema>"""

VALID_XML = "<person><name>Karl</name><age>30</age></person>"
INVALID_XML = "<person><name>Karl</name><age>thirty</age><extra/></person>"


@pytest.fixture
def client() -> TestClient:
    return TestClient(app)


def _load_xsd(client: TestClient) -> str:
    r = client.post("/api/xsd/text", json={"content": XSD, "filename": "person.xsd"})
    assert r.status_code == 200, r.text
    return r.json()["xsd_id"]


def _load_xml(client: TestClient, content: str, filename: str) -> str:
    r = client.post("/api/xml/text", json={"content": content, "filename": filename})
    assert r.status_code == 200, r.text
    return r.json()["xml_id"]


def test_health(client: TestClient) -> None:
    assert client.get("/api/health").json()["status"] == "ok"


def test_xml_tree_structure(client: TestClient) -> None:
    r = client.post("/api/xml/text", json={"content": VALID_XML, "filename": "v.xml"})
    body = r.json()
    assert body["root"]["tag"] == "person"
    assert body["node_count"] == 3
    children = [c["tag"] for c in body["root"]["children"]]
    assert children == ["name", "age"]


def test_valid_document(client: TestClient) -> None:
    xsd_id = _load_xsd(client)
    xml_id = _load_xml(client, VALID_XML, "v.xml")
    r = client.post("/api/validate", json={"xml_id": xml_id, "xsd_id": xsd_id})
    body = r.json()
    assert body["is_valid"] is True
    assert body["errors"] == []


def test_invalid_document_maps_errors_to_nodes(client: TestClient) -> None:
    xsd_id = _load_xsd(client)
    xml_id = _load_xml(client, INVALID_XML, "i.xml")
    r = client.post("/api/validate", json={"xml_id": xml_id, "xsd_id": xsd_id})
    body = r.json()
    assert body["is_valid"] is False
    assert len(body["errors"]) == 2
    # Every error must resolve to a tree node id for inline highlighting.
    assert all(e["node_id"] is not None for e in body["errors"])


def test_excel_report_is_readable(client: TestClient) -> None:
    xsd_id = _load_xsd(client)
    xml_id = _load_xml(client, INVALID_XML, "i.xml")
    vid = client.post(
        "/api/validate", json={"xml_id": xml_id, "xsd_id": xsd_id}
    ).json()["validation_id"]

    r = client.get(f"/api/validate/{vid}/excel")
    assert r.status_code == 200
    assert r.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument"
    )
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb.active
    # Meta block (5 rows) + blank + header + 2 error rows.
    assert ws.max_row >= 9


def test_not_well_formed_xml_rejected(client: TestClient) -> None:
    r = client.post("/api/xml/text", json={"content": "<a><b></a>", "filename": "x.xml"})
    assert r.status_code == 400
    assert "well-formed" in r.json()["detail"]


def test_invalid_xsd_rejected(client: TestClient) -> None:
    r = client.post("/api/xsd/text", json={"content": "<nope/>", "filename": "x.xsd"})
    assert r.status_code == 422


def test_unknown_ids_404(client: TestClient) -> None:
    r = client.post("/api/validate", json={"xml_id": "deadbeef", "xsd_id": "deadbeef"})
    assert r.status_code == 404


def test_dtd_rejected(client: TestClient) -> None:
    bomb = '<?xml version="1.0"?><!DOCTYPE x [<!ENTITY a "b">]><x/>'
    r = client.post("/api/xml/text", json={"content": bomb, "filename": "x.xml"})
    assert r.status_code == 400


# --- Feedback ---------------------------------------------------------------


def test_feedback_without_db_is_logged_and_accepted(client: TestClient, caplog) -> None:
    with caplog.at_level("WARNING", logger="app.api.feedback"):
        r = client.post("/api/feedback", json={"message": "Nice tool!", "page": "/"})
    assert r.status_code == 204, r.text
    assert any("user feedback" in rec.getMessage() for rec in caplog.records)


def test_feedback_honeypot_is_dropped_silently(client: TestClient, caplog) -> None:
    with caplog.at_level("WARNING", logger="app.api.feedback"):
        r = client.post("/api/feedback", json={"message": "spam", "website": "http://x"})
    assert r.status_code == 204
    assert not any("user feedback" in rec.getMessage() for rec in caplog.records)


def test_feedback_rejects_blank_and_bad_email(client: TestClient) -> None:
    assert client.post("/api/feedback", json={"message": "   "}).status_code == 422
    assert client.post("/api/feedback", json={"message": "x", "email": "nope"}).status_code == 422


def test_feedback_uses_store_when_configured(client: TestClient) -> None:
    import asyncio

    from app.feedback_store import FeedbackRow, FeedbackStore

    saved: list[tuple] = []

    class _Cursor:
        async def __aenter__(self):
            return self

        async def __aexit__(self, *a):
            return False

        async def execute(self, sql, row):
            saved.append(row)

    class _Conn:
        def cursor(self):
            return _Cursor()

        async def commit(self):
            pass

        async def close(self):
            pass

    async def connect(dsn, **kw):
        return _Conn()

    store = FeedbackStore("postgresql://x/y", connect=connect)
    asyncio.run(store.save(FeedbackRow(message="hi", xml_name="a.xml")))
    assert saved and saved[0][0] == "hi" and saved[0][3] == "a.xml"
